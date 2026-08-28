"""Task 9: staged publication must be single-writer and failure-atomic."""

from __future__ import annotations

from dataclasses import replace
from hashlib import sha256
import json
import os
from pathlib import Path
import shutil
import tempfile
import unittest
from unittest.mock import patch

from openpyxl import load_workbook

from skill_maintainer.catalog import Catalog, CatalogRow, CatalogSnapshot, diff_catalog
from skill_maintainer.ledger import LedgerStore
from skill_maintainer.ledger_schema import CURRENT_SKILL_COLUMNS
from skill_maintainer.locking import LockUnavailable, SingleWriterLock
from skill_maintainer.review import DerivedFields, ObservedFacts, ProjectJudgments, ReviewDecision, build_review_packet
from skill_maintainer.runner import (
    CoordinatorError,
    RunCoordinator,
    RunRequest,
    SourceRun,
)
from skill_maintainer.sources.base import SourceRequestEvent
from skill_maintainer.snapshots import SnapshotCandidate, build_snapshot


class RunnerTestCase(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory()
        self.addCleanup(self.temporary.cleanup)
        self.root = Path(self.temporary.name) / "项目 根目录"
        self.root.mkdir()
        self.settings = self.root / "workflow-settings.toml"
        self.settings.write_text(
            """config_version = 1

[workflow]
enabled = false
timezone = \"Asia/Shanghai\"

[schedule]
mode = \"manual\"
start_time = \"22:00\"
weekdays = [\"Monday\"]
interval_days = 1
day_of_month = 1

[research]
incremental_search = true
full_recheck_interval_days = 7
check_existing_skill_updates = true
include_generic_skills = false

[delivery]
generate_word = true
generate_excel = true
only_refresh_affected_classes = true
notify_on_no_change = false
""",
            encoding="utf-8",
        )
        ledger = self.root / "ledger" / "Skills主台账.xlsx"
        LedgerStore.create(ledger)
        self.request = RunRequest(settings_path=self.settings, catalog_loader=lambda: object())

    def coordinator(self, **overrides):
        options = {
            "root": self.root,
            "discover": lambda request, staging: (
                SourceRun("SkillHub", "partial"), SourceRun("ClawHub", "failed"),
                SourceRun("GitHub", "failed"), SourceRun("Hugging Face Spaces", "failed"),
            ),
            "report_builder": lambda prepared, staging: (),
            "office_verifier": lambda prepared, artifacts: True,
        }
        options.update(overrides)
        return RunCoordinator(**options)

    def production_hashes(self) -> dict[str, str]:
        result = {}
        for path in sorted(item for item in self.root.rglob("*") if item.is_file() and ".runtime" not in item.parts):
            result[str(path.relative_to(self.root))] = sha256(path.read_bytes()).hexdigest()
        return result

    @staticmethod
    def formal_row(stable_id: str, version: str, content_hash: str) -> dict[str, object]:
        row = {column: "已核验" for column in CURRENT_SKILL_COLUMNS}
        row.update({
            "内部标识": stable_id, "Skill名称": stable_id, "规范名称": stable_id, "入库层级": "正式",
            "来源平台": "GitHub", "发现地址": "https://github.com/example/skill", "Canonical source": "https://github.com/example/skill",
            "上游项目地址": "https://github.com/example/skill", "Skill入口路径": "SKILL.md", "固定版本": version,
            "固定版本内容指纹": content_hash, "许可证": "MIT", "外部联网/API 调用": "否", "远程服务端点": "",
            "安全等级": "SA", "验证状态": "全部通过（未实测）", "质量评分": 3,
        })
        return row

    @staticmethod
    def report_scope_catalog() -> Catalog:
        return Catalog((
            CatalogRow("08", "工学", "0809", "计算机类", "080901", "计算机科学与技术"),
            CatalogRow("11", "军事学", "1101", "军事类", "110101", "军事专业"),
            CatalogRow("14", "交叉学科", None, None, "140101", "集成电路科学与工程"),
        ))

    def assert_report_mapping_rejected_before_output(self, stable_id: str, code: str, name: str) -> None:
        from skill_maintainer.reports import ReportBuildError, make_project_report_builder

        production = self.root / "ledger" / "Skills主台账.xlsx"
        ledger = LedgerStore.load(production)
        ledger.append_rows("专业任务映射", [{
            "映射标识": f"MAP-{stable_id}", "内部标识": stable_id,
            "专业代码": code, "专业名称": name, "专业任务": "不应纳入的任务",
            "输入": "数据", "输出": "报告", "适用理由": "不应纳入", "使用限制": "禁止", "相关度": "高",
            "专业别名": name, "核心课程": name, "研究方法": name, "工作任务": name,
            "成果或数据对象": name, "软件/数据库/流程": name,
        }])
        seeded = self.root / "ledger" / f"{stable_id}-seed.xlsx"
        ledger.save_staged(seeded)
        ledger.workbook.close()
        shutil.copyfile(seeded, production)
        seeded.unlink()

        packet, decision = self.report_review(stable_id, f"https://github.com/example/{stable_id.lower()}")
        coordinator = self.coordinator(report_builder=make_project_report_builder(self.root))
        request = replace(
            self.request,
            requested_run_id=f"run-reject-{stable_id.lower()}",
            catalog_loader=self.report_scope_catalog,
            review_packets={stable_id: packet},
        )
        prepared = coordinator.prepare(request)
        coordinator.apply_reviews(prepared, (decision,))
        try:
            with self.assertRaisesRegex(ReportBuildError, f"{stable_id}|专业任务映射"):
                coordinator.report_builder(prepared, prepared.staging_dir)
            delivery = prepared.staging_dir / "deliveries"
            self.assertTrue(not delivery.exists() or not any(delivery.iterdir()))
        finally:
            coordinator.abandon(prepared)

    def report_review(self, stable_id: str, canonical_source: str):
        candidate_root = self.root / f"{stable_id}-candidate"
        candidate_root.mkdir()
        (candidate_root / "SKILL.md").write_text(f"# {stable_id}", encoding="utf-8")
        version = "c" * 40
        evidence = (f"https://evidence.example/{stable_id}",)
        snapshot = build_snapshot(
            SnapshotCandidate(stable_id, version, candidate_root, evidence),
            self.root / f"{stable_id}-snapshot",
        )
        packet = build_review_packet({
            "id": stable_id, "canonical_source": canonical_source,
            "license": "MIT", "security_grade": "SA",
        }, snapshot)
        proposed = self.formal_row(stable_id, version, snapshot.fixed_content_hash)
        proposed.update({
            "发现地址": canonical_source, "Canonical source": canonical_source,
            "上游项目地址": canonical_source, "验证证据位置": "；".join(evidence),
            "本地专业软件或运行时依赖": "无", "本地脚本/插件接口": "不使用", "质量评分": 2,
        })
        decision = ReviewDecision(
            ObservedFacts(version, True, True, "MIT", canonical_source, evidence, "否", (), "无", "不使用", "SA", "全部通过（未实测）"),
            ProjectJudgments("正式推荐", True, True, 4, (True,)),
            DerivedFields(quality_score=2, ledger_row=proposed), candidate_id=stable_id,
        )
        return packet, decision

    def test_live_lock_blocks_second_holder_and_stale_diagnostic_is_recoverable(self):
        lock_path = self.root / "runtime" / "writer.lock"
        first = SingleWriterLock(lock_path)
        self.assertTrue(first.acquire())
        self.addCleanup(first.release)
        with self.assertRaises(LockUnavailable):
            SingleWriterLock(lock_path).acquire()
        first.release()
        lock_path.write_text("pid=not-an-owner", encoding="utf-8")
        recovered = SingleWriterLock(lock_path)
        self.assertTrue(recovered.acquire())
        recovered.release()

    def test_lock_and_staging_link_or_traversal_are_rejected(self):
        link = self.root / "runtime-link"
        try:
            link.symlink_to(self.root, target_is_directory=True)
        except (NotImplementedError, OSError) as exc:
            self.skipTest(f"test filesystem cannot create a link: {exc}")
        with self.assertRaises(ValueError):
            SingleWriterLock(link / "writer.lock").acquire()
        coordinator = self.coordinator()
        with self.assertRaises(ValueError):
            coordinator.prepare(replace(self.request, requested_run_id="../escape"))

    def test_prepare_and_repeated_identical_apply_are_staging_only_and_idempotent(self):
        coordinator = self.coordinator()
        before = self.production_hashes()
        prepared = coordinator.prepare(self.request)
        self.assertEqual(before, self.production_hashes())
        self.assertTrue(prepared.staging_ledger.exists())
        first = coordinator.apply_reviews(prepared, ())
        second = coordinator.apply_reviews(prepared, ())
        self.assertEqual(first, second)
        self.assertEqual(before, self.production_hashes())
        coordinator.abandon(prepared)

    def test_real_report_adapter_uses_prepared_catalog_sources_and_both_ledgers_in_staging(self):
        if not os.environ.get("SKILL_MAINTAINER_NODE") or not os.environ.get("SKILL_MAINTAINER_NODE_MODULES"):
            self.skipTest("report integration requires caller-supplied Node runtime")
        from skill_maintainer.reports import make_project_report_builder

        production = self.root / "ledger" / "Skills主台账.xlsx"
        ledger = LedgerStore.load(production)
        ledger.append_rows("当前Skill", [self.formal_row("EXISTING-REPORT-1", "v1", "a" * 64)])
        ledger.append_rows("专业任务映射", [
            {
                "映射标识": "MAP-REPORT-1", "内部标识": "EXISTING-REPORT-1",
                "专业代码": "0809", "专业名称": "计算机类", "专业任务": "课程分析",
                "输入": "课程表", "输出": "报告", "适用理由": "直接相关", "使用限制": "脱敏", "相关度": "高",
                "专业别名": "计算机", "核心课程": "程序设计", "研究方法": "数据分析", "工作任务": "课程治理",
                "成果或数据对象": "课程表", "软件/数据库/流程": "Python",
            },
            {
                "映射标识": "MAP-REPORT-NEW", "内部标识": "REPORT-NEW-1",
                "专业代码": "0809", "专业名称": "计算机类", "专业任务": "课程分析",
                "输入": "课程表", "输出": "报告", "适用理由": "直接相关", "使用限制": "脱敏", "相关度": "高",
                "专业别名": "计算机", "核心课程": "程序设计", "研究方法": "数据分析", "工作任务": "课程治理",
                "成果或数据对象": "课程表", "软件/数据库/流程": "Python",
            },
        ])
        seeded = self.root / "ledger" / "seed-report.xlsx"
        ledger.save_staged(seeded)
        ledger.workbook.close()
        shutil.copyfile(seeded, production)
        seeded.unlink()

        old_rows = (CatalogRow("08", "工学", "0801", "力学类", "080101", "理论与应用力学"),)
        new_rows = (*old_rows, CatalogRow("08", "工学", "0809", "计算机类", "080901", "计算机科学与技术"))
        catalog = Catalog(old_rows, staged_snapshot=CatalogSnapshot(new_rows, "b" * 64), staged_diff=diff_catalog(old_rows, new_rows))
        request_event = SourceRequestEvent(
            "GitHub", "query-0809", "https://api.github.com/search/repositories?q=campus", 2,
            200, 3, "d" * 64, None, last_page=True,
            evidence_path=self.root / "evidence" / "github-page-2.json", completed=True,
        )
        sources = (
            SourceRun("SkillHub", "partial"), SourceRun("ClawHub", "failed"),
            SourceRun("GitHub", "complete", request_events=(request_event,)), SourceRun("Hugging Face Spaces", "failed"),
        )
        packet, decision = self.report_review("REPORT-NEW-1", "https://github.com/example/report-new")
        coordinator = self.coordinator(
            discover=lambda request, staging: sources,
            report_builder=make_project_report_builder(self.root),
        )
        request = replace(self.request, catalog_loader=lambda: catalog, review_packets={"REPORT-NEW-1": packet})
        prepared = coordinator.prepare(request)
        self.assertIs(prepared.catalog_snapshot, catalog)
        summary = coordinator.finalize(prepared, coordinator.apply_reviews(prepared, (decision,)))
        generation = summary.output_generation
        self.assertIsNotNone(generation)
        self.assertTrue((generation / "维护日报.docx").is_file())
        daily = load_workbook(generation / "维护日报.xlsx", data_only=False)
        self.addCleanup(daily.close)
        audit = daily["来源请求审计"]
        audit_text = "\n".join(str(cell.value or "") for row in audit.iter_rows() for cell in row)
        for expected in ("GitHub", request_event.url, "query-0809", "2", "200", "3", "d" * 64, str(request_event.evidence_path), "是", "未记录"):
            self.assertIn(expected, audit_text)
        self.assertNotIn("__run__", audit_text)
        self.assertEqual(audit["B2"].hyperlink.target, request_event.url)
        self.assertEqual(daily["新增正式推荐"]["A2"].value, "REPORT-NEW-1")
        scope_books = tuple(generation.glob("受影响专业类/*/专业类Skill清单.xlsx"))
        self.assertEqual(len(scope_books), 1)
        scope_book = load_workbook(scope_books[0], data_only=False)
        self.addCleanup(scope_book.close)
        self.assertEqual(scope_book["新增正式推荐"]["A2"].value, "EXISTING-REPORT-1")

    def test_report_adapter_rejects_unmapped_new_formal_before_any_output(self):
        if not os.environ.get("SKILL_MAINTAINER_NODE") or not os.environ.get("SKILL_MAINTAINER_NODE_MODULES"):
            self.skipTest("report integration requires caller-supplied Node runtime")
        from skill_maintainer.reports import ReportBuildError, make_project_report_builder

        packet, decision = self.report_review("UNMAPPED-REPORT-1", "https://github.com/example/unmapped-report")
        coordinator = self.coordinator(report_builder=make_project_report_builder(self.root))
        request = replace(
            self.request,
            catalog_loader=self.report_scope_catalog,
            review_packets={"UNMAPPED-REPORT-1": packet},
        )
        prepared = coordinator.prepare(request)
        coordinator.apply_reviews(prepared, (decision,))
        try:
            with self.assertRaisesRegex(ReportBuildError, "UNMAPPED-REPORT-1|专业任务映射"):
                coordinator.report_builder(prepared, prepared.staging_dir)
            delivery = prepared.staging_dir / "deliveries"
            self.assertTrue(not delivery.exists() or not any(delivery.iterdir()))
        finally:
            coordinator.abandon(prepared)

    def test_report_adapter_rejects_military_mapping_before_any_output(self):
        if not os.environ.get("SKILL_MAINTAINER_NODE") or not os.environ.get("SKILL_MAINTAINER_NODE_MODULES"):
            self.skipTest("report integration requires caller-supplied Node runtime")
        from skill_maintainer.reports import ReportBuildError, make_project_report_builder

        production = self.root / "ledger" / "Skills主台账.xlsx"
        ledger = LedgerStore.load(production)
        ledger.append_rows("专业任务映射", [{
            "映射标识": "MAP-MILITARY-1", "内部标识": "MILITARY-REPORT-1",
            "专业代码": "1101", "专业名称": "军事类", "专业任务": "军事任务",
            "输入": "数据", "输出": "报告", "适用理由": "不应纳入", "使用限制": "禁止", "相关度": "高",
            "专业别名": "军事", "核心课程": "军事", "研究方法": "军事", "工作任务": "军事",
            "成果或数据对象": "军事", "软件/数据库/流程": "军事",
        }])
        seeded = self.root / "ledger" / "military-seed.xlsx"
        ledger.save_staged(seeded)
        ledger.workbook.close()
        shutil.copyfile(seeded, production)
        seeded.unlink()
        packet, decision = self.report_review("MILITARY-REPORT-1", "https://github.com/example/military-report")
        coordinator = self.coordinator(report_builder=make_project_report_builder(self.root))
        request = replace(
            self.request,
            catalog_loader=self.report_scope_catalog,
            review_packets={"MILITARY-REPORT-1": packet},
        )
        prepared = coordinator.prepare(request)
        coordinator.apply_reviews(prepared, (decision,))
        try:
            with self.assertRaisesRegex(ReportBuildError, "MILITARY-REPORT-1|军事|专业任务映射"):
                coordinator.report_builder(prepared, prepared.staging_dir)
            delivery = prepared.staging_dir / "deliveries"
            self.assertTrue(not delivery.exists() or not any(delivery.iterdir()))
        finally:
            coordinator.abandon(prepared)

    def test_report_adapter_rejects_malformed_prefix_mapping_before_any_output(self):
        from skill_maintainer.reports import ReportBuildError, make_project_report_builder

        production = self.root / "ledger" / "Skills主台账.xlsx"
        ledger = LedgerStore.load(production)
        ledger.append_rows("专业任务映射", [{
            "映射标识": "MAP-MALFORMED-1", "内部标识": "MALFORMED-REPORT-1",
            "专业代码": "08evil", "专业名称": "伪造类", "专业任务": "伪造任务",
            "输入": "数据", "输出": "报告", "适用理由": "不应纳入", "使用限制": "禁止", "相关度": "高",
            "专业别名": "伪造", "核心课程": "伪造", "研究方法": "伪造", "工作任务": "伪造",
            "成果或数据对象": "伪造", "软件/数据库/流程": "伪造",
        }])
        seeded = self.root / "ledger" / "malformed-seed.xlsx"
        ledger.save_staged(seeded)
        ledger.workbook.close()
        shutil.copyfile(seeded, production)
        seeded.unlink()

        packet, decision = self.report_review("MALFORMED-REPORT-1", "https://github.com/example/malformed-report")
        coordinator = self.coordinator(report_builder=make_project_report_builder(self.root))
        request = replace(
            self.request,
            catalog_loader=self.report_scope_catalog,
            review_packets={"MALFORMED-REPORT-1": packet},
        )
        prepared = coordinator.prepare(request)
        coordinator.apply_reviews(prepared, (decision,))
        try:
            with self.assertRaisesRegex(ReportBuildError, "MALFORMED-REPORT-1|专业任务映射"):
                coordinator.report_builder(prepared, prepared.staging_dir)
            delivery = prepared.staging_dir / "deliveries"
            self.assertTrue(not delivery.exists() or not any(delivery.iterdir()))
        finally:
            coordinator.abandon(prepared)

    def test_report_adapter_rejects_blank_code_even_when_name_starts_with_approved_code(self):
        if not os.environ.get("SKILL_MAINTAINER_NODE") or not os.environ.get("SKILL_MAINTAINER_NODE_MODULES"):
            self.skipTest("report integration requires caller-supplied Node runtime")
        cases = (
            ("BLANK-ORDINARY-REPORT-1", "0809 任意自由文本"),
            ("BLANK-MILITARY-REPORT-1", "0809 军事自由文本"),
        )
        for stable_id, name in cases:
            with self.subTest(name=name):
                self.assert_report_mapping_rejected_before_output(stable_id, "", name)

    def test_all_failure_points_leave_production_bytes_identical(self):
        points = ("after_discovery", "after_review", "report", "office", "reopen", "before_publish")
        for point in points:
            with self.subTest(point=point):
                baseline = self.production_hashes()
                coordinator = self.coordinator(fail_at=point)
                if point == "after_discovery":
                    with self.assertRaises(RuntimeError):
                        coordinator.prepare(self.request)
                    self.assertEqual(baseline, self.production_hashes())
                    continue
                prepared = coordinator.prepare(self.request)
                if point == "after_review":
                    with self.assertRaises(RuntimeError):
                        coordinator.apply_reviews(prepared, ())
                else:
                    reviews = coordinator.apply_reviews(prepared, ())
                    with self.assertRaises(RuntimeError):
                        coordinator.finalize(prepared, reviews)
                self.assertEqual(baseline, self.production_hashes())

    def test_partial_sources_advance_only_complete_watermarks(self):
        batches = (
            SourceRun("SkillHub", "complete", watermark="skillhub-2"),
            SourceRun("ClawHub", "partial", watermark="clawhub-2"),
            SourceRun("GitHub", "failed", watermark="github-2"),
            SourceRun("Hugging Face Spaces", "complete", watermark="hf-2"),
        )
        coordinator = self.coordinator(discover=lambda request, staging: batches)
        prepared = coordinator.prepare(self.request)
        summary = coordinator.finalize(prepared, coordinator.apply_reviews(prepared, ()))
        self.assertEqual(summary.source_statuses, {item.platform: item.status for item in batches})
        rows = LedgerStore.load(self.root / "ledger" / "Skills主台账.xlsx").rows("来源水位")
        self.assertEqual({row["来源平台"] for row in rows}, {"SkillHub", "Hugging Face Spaces"})

    def test_all_sources_failed_writes_only_staging_failure_report(self):
        failed = tuple(SourceRun(name, "failed") for name in ("SkillHub", "ClawHub", "GitHub", "Hugging Face Spaces"))
        before = self.production_hashes()
        coordinator = self.coordinator(discover=lambda request, staging: failed)
        prepared = coordinator.prepare(self.request)
        summary = coordinator.finalize(prepared, coordinator.apply_reviews(prepared, ()))
        self.assertTrue(summary.blocked)
        self.assertTrue((prepared.staging_dir / "failure-report.json").is_file())
        self.assertEqual(before, self.production_hashes())

    def test_tamper_stale_duplicate_finalize_and_cross_coordinator_are_rejected(self):
        coordinator = self.coordinator()
        prepared = coordinator.prepare(self.request)
        reviews = coordinator.apply_reviews(prepared, ())
        prepared.staging_ledger.write_bytes(prepared.staging_ledger.read_bytes() + b"tamper")
        with self.assertRaises(CoordinatorError):
            coordinator.finalize(prepared, reviews)

        fresh = self.coordinator()
        prepared = fresh.prepare(self.request)
        with self.assertRaises(CoordinatorError):
            self.coordinator().apply_reviews(prepared, ())
        reviews = fresh.apply_reviews(prepared, ())
        fresh.finalize(prepared, reviews)
        with self.assertRaises(CoordinatorError):
            fresh.finalize(prepared, reviews)

    def test_config_change_after_prepare_and_mid_publish_failure_cannot_expose_a_new_generation(self):
        coordinator = self.coordinator()
        prepared = coordinator.prepare(self.request)
        self.settings.write_text(self.settings.read_text(encoding="utf-8") + "\n", encoding="utf-8")
        with self.assertRaises(CoordinatorError):
            coordinator.apply_reviews(prepared, ())

        # A publication transaction may create a staged generation, but it must restore the
        # existing ledger and pointer if interrupted before its one visible pointer update.
        (self.root / "output").mkdir()
        (self.root / "output" / "current-generation.json").write_text('{"run_id":"old"}', encoding="utf-8")
        baseline = self.production_hashes()

        def write_delivery(prepared, staging):
            delivery = staging / "deliveries"
            delivery.mkdir()
            report = delivery / "report.txt"
            report.write_text("new", encoding="utf-8")
            return (report,)

        coordinator = self.coordinator(report_builder=write_delivery, fail_at="before_commit")
        prepared = coordinator.prepare(self.request)
        reviews = coordinator.apply_reviews(prepared, ())
        with self.assertRaises(RuntimeError):
            coordinator.finalize(prepared, reviews)
        after = self.production_hashes()
        self.assertEqual(baseline["ledger\\Skills主台账.xlsx"], after["ledger\\Skills主台账.xlsx"])
        self.assertEqual(baseline["output\\current-generation.json"], after["output\\current-generation.json"])

    def test_failure_before_single_ledger_commit_preserves_old_ledger_and_generation_pointer(self):
        (self.root / "output").mkdir()
        pointer = self.root / "output" / "current-generation.json"
        pointer.write_text('{"run_id":"old"}', encoding="utf-8")
        baseline = self.production_hashes()

        def write_delivery(prepared, staging):
            delivery = staging / "deliveries"
            delivery.mkdir()
            report = delivery / "report.txt"
            report.write_text("new", encoding="utf-8")
            return (report,)

        coordinator = self.coordinator(report_builder=write_delivery, fail_at="before_commit")
        prepared = coordinator.prepare(self.request)
        reviews = coordinator.apply_reviews(prepared, ())
        with self.assertRaises(RuntimeError):
            coordinator.finalize(prepared, reviews)
        after = self.production_hashes()
        self.assertEqual(baseline["ledger\\Skills主台账.xlsx"], after["ledger\\Skills主台账.xlsx"])
        self.assertEqual(baseline["output\\current-generation.json"], after["output\\current-generation.json"])

    def test_catalog_change_after_prepare_releases_the_registry_and_writer_lock(self):
        catalog = {"revision": "one"}
        request = replace(self.request, catalog_loader=lambda: dict(catalog))
        coordinator = self.coordinator()
        prepared = coordinator.prepare(request)
        catalog["revision"] = "two"
        with self.assertRaises(CoordinatorError):
            coordinator.apply_reviews(prepared, ())
        # The failure finalizer must release its OS lock: a fresh writer may acquire it.
        successor = self.coordinator()
        next_run = successor.prepare(self.request)
        successor.abandon(next_run)

    def test_task7_packets_are_applied_only_to_staging_and_receipts_are_consumed_on_success(self):
        candidate_root = self.root / "review-candidate"
        candidate_root.mkdir()
        (candidate_root / "SKILL.md").write_text("# static evidence", encoding="utf-8")
        version = "a" * 40
        evidence = ("https://evidence.example/commit",)
        snapshot = build_snapshot(SnapshotCandidate("candidate-1", version, candidate_root, evidence), self.root / "snapshot")
        packet = build_review_packet(
            {"id": "candidate-1", "canonical_source": "https://github.com/example/skill", "license": "MIT", "security_grade": "SA"},
            snapshot,
        )
        decision = ReviewDecision(
            ObservedFacts(version, True, True, "MIT", "https://github.com/example/skill", evidence, "否", (), "无", "不使用", "SA", "全部通过（未实测）"),
            ProjectJudgments("正式推荐", True, True, 4, (True,)),
            candidate_id="candidate-1",
        )
        coordinator = self.coordinator()
        prepared = coordinator.prepare(replace(self.request, review_packets={"candidate-1": packet}))
        reviews = coordinator.apply_reviews(prepared, (decision,))
        self.assertEqual(reviews.applied_count, 1)
        coordinator.finalize(prepared, reviews)

    def test_source_contract_requires_exact_platforms_and_binds_candidate_evidence(self):
        evidence = self.root / "evidence.json"
        evidence.write_text('{"safe":true}', encoding="utf-8")
        complete = SourceRun("SkillHub", "complete", candidates=({"id": "new"},), evidence_files=(evidence,))
        others = tuple(SourceRun(name, "failed") for name in ("ClawHub", "GitHub", "Hugging Face Spaces"))
        coordinator = self.coordinator(discover=lambda request, staging: (complete, *others))
        prepared = coordinator.prepare(self.request)
        evidence.write_text('{"tampered":true}', encoding="utf-8")
        with self.assertRaises(CoordinatorError):
            coordinator.apply_reviews(prepared, ())

        incomplete = self.coordinator(discover=lambda request, staging: (SourceRun("SkillHub", "failed"),))
        with self.assertRaises(CoordinatorError):
            incomplete.prepare(self.request)

    def test_discovered_skills_and_manual_review_are_retained_without_dangling_aliases(self):
        evidence = self.root / "github-evidence.json"
        evidence.write_text("{}", encoding="utf-8")
        candidate = {
            "id": "candidate-new", "native_id": "candidate-new", "platform": "GitHub",
            "discovery_url": "https://github.com/example/new", "canonical_source": "https://github.com/example/new",
            "upstream_identity": "example/new", "entry_path": "SKILL.md", "content_hash": "a" * 64,
        }
        batches = (
            SourceRun("SkillHub", "failed"), SourceRun("ClawHub", "failed"),
            SourceRun("GitHub", "complete", candidates=(candidate,), evidence_files=(evidence,)),
            SourceRun("Hugging Face Spaces", "failed"),
        )
        coordinator = self.coordinator(discover=lambda request, staging: batches)
        prepared = coordinator.prepare(self.request)
        summary = coordinator.apply_reviews(prepared, ())
        coordinator.finalize(prepared, summary)
        ledger = LedgerStore.load(self.root / "ledger" / "Skills主台账.xlsx")
        observations = ledger.rows("候选观察")
        aliases = ledger.rows("来源别名")
        self.assertEqual(ledger.rows("当前Skill"), [])
        self.assertTrue(any(row["Canonical source"] == candidate["canonical_source"] for row in observations))
        current_ids = {row["内部标识"] for row in ledger.rows("当前Skill")}
        self.assertTrue(all(row["内部标识"] in current_ids for row in aliases))

    def test_successful_ledger_commit_is_the_only_authority_for_complete_generation(self):
        platforms = tuple(SourceRun(name, "failed") for name in ("SkillHub", "ClawHub", "GitHub", "Hugging Face Spaces"))
        # Use one partial source to avoid the all-failed business block while proving no pointer is written.
        platforms = (SourceRun("SkillHub", "partial"), *platforms[1:])
        coordinator = self.coordinator(discover=lambda request, staging: platforms)
        prepared = coordinator.prepare(self.request)
        summary = coordinator.finalize(prepared, coordinator.apply_reviews(prepared, ()))
        self.assertIsNotNone(summary.output_generation)
        self.assertFalse((self.root / "output" / "current-generation.json").exists())
        rows = LedgerStore.load(self.root / "ledger" / "Skills主台账.xlsx").rows("运行记录")
        self.assertEqual(rows[-1]["状态"], "成功")
        self.assertIn(summary.run_id, rows[-1]["摘要"])

    def test_source_contract_rejects_unknown_duplicate_and_missing_platforms(self):
        exact = ("SkillHub", "ClawHub", "GitHub", "Hugging Face Spaces")
        for broken in (
            tuple(SourceRun(name, "failed") for name in exact[:-1]),
            (SourceRun("SkillHub", "failed"), SourceRun("SkillHub", "failed"), *tuple(SourceRun(name, "failed") for name in exact[1:])),
            (SourceRun("Unknown", "failed"), *tuple(SourceRun(name, "failed") for name in exact[1:])),
        ):
            with self.subTest(broken=tuple(item.platform for item in broken)):
                coordinator = self.coordinator(discover=lambda request, staging, result=broken: result)
                with self.assertRaises(CoordinatorError):
                    coordinator.prepare(self.request)

    def test_base_exception_before_commit_releases_lock_and_never_writes_a_pointer(self):
        (self.root / "output").mkdir()
        old_pointer = self.root / "output" / "current-generation.json"
        old_pointer.write_text('{"run_id":"old"}', encoding="utf-8")
        before = self.production_hashes()
        source_runs = (SourceRun("SkillHub", "partial"), *tuple(SourceRun(name, "failed") for name in ("ClawHub", "GitHub", "Hugging Face Spaces")))
        coordinator = self.coordinator(discover=lambda request, staging: source_runs, before_publish=lambda prepared: (_ for _ in ()).throw(SystemExit("crash")))
        prepared = coordinator.prepare(self.request)
        with self.assertRaises(SystemExit):
            coordinator.finalize(prepared, coordinator.apply_reviews(prepared, ()))
        after = self.production_hashes()
        self.assertEqual(before["ledger\\Skills主台账.xlsx"], after["ledger\\Skills主台账.xlsx"])
        self.assertEqual(before["output\\current-generation.json"], after["output\\current-generation.json"])
        successor = self.coordinator()
        next_run = successor.prepare(self.request)
        successor.abandon(next_run)

    def test_generation_link_is_rejected_before_ledger_commit_and_state_is_popped(self):
        outside = self.root / "outside"
        outside.mkdir()
        source_runs = (SourceRun("SkillHub", "partial"), *tuple(SourceRun(name, "failed") for name in ("ClawHub", "GitHub", "Hugging Face Spaces")))

        def unsafe_delivery(prepared, staging):
            delivery = staging / "deliveries"
            delivery.mkdir()
            (delivery / "report.txt").write_text("safe", encoding="utf-8")
            try:
                (delivery / "junction").symlink_to(outside, target_is_directory=True)
            except OSError as error:
                self.skipTest(f"test filesystem cannot create a link: {error}")
            return (delivery / "report.txt",)

        coordinator = self.coordinator(discover=lambda request, staging: source_runs, report_builder=unsafe_delivery)
        prepared = coordinator.prepare(self.request)
        with self.assertRaises(CoordinatorError):
            coordinator.finalize(prepared, coordinator.apply_reviews(prepared, ()))
        self.assertNotIn(prepared.run_id, coordinator._states)

    def test_existing_version_review_stays_out_of_current_skill_until_finalize_transaction(self):
        old_hash, new_hash = "b" * 64, "c" * 64
        production = self.root / "ledger" / "Skills主台账.xlsx"
        base = LedgerStore.load(production)
        base.append_rows("当前Skill", [self.formal_row("EXISTING-1", "v1", old_hash)])
        base.save_staged(self.root / "ledger" / "seed.xlsx")
        shutil.copyfile(self.root / "ledger" / "seed.xlsx", production)

        candidate_root = self.root / "version-candidate"
        candidate_root.mkdir()
        (candidate_root / "SKILL.md").write_text("new version", encoding="utf-8")
        version = "d" * 40
        evidence = ("https://evidence.example/version",)
        snapshot = build_snapshot(SnapshotCandidate("EXISTING-1", version, candidate_root, evidence), self.root / "version-snapshot")
        packet = build_review_packet({"id": "EXISTING-1", "canonical_source": "https://github.com/example/skill", "license": "MIT", "security_grade": "SA"}, snapshot)
        proposed = self.formal_row("EXISTING-1", version, snapshot.fixed_content_hash)
        proposed.update({"验证证据位置": "；".join(evidence), "本地专业软件或运行时依赖": "无", "本地脚本/插件接口": "不使用", "质量评分": 2})
        decision = ReviewDecision(
            ObservedFacts(version, True, True, "MIT", "https://github.com/example/skill", evidence, "否", (), "无", "不使用", "SA", "全部通过（未实测）"),
            ProjectJudgments("正式推荐", True, True, 4, (True,)), DerivedFields(quality_score=2, ledger_row=proposed), "EXISTING-1",
        )
        coordinator = self.coordinator()
        prepared = coordinator.prepare(replace(self.request, review_packets={"EXISTING-1": packet}))
        reviews = coordinator.apply_reviews(prepared, (decision,))
        staged_before_finalize = LedgerStore.load(prepared.staging_ledger).rows("当前Skill")[0]
        self.assertEqual(staged_before_finalize["固定版本"], "v1")
        coordinator.finalize(prepared, reviews)
        current = LedgerStore.load(production).rows("当前Skill")[0]
        self.assertEqual(current["固定版本"], version)

    def test_one_project_terminal_cleanup_does_not_clear_another_projects_task7_packet(self):
        second_root = self.root.parent / "second-project"
        second_root.mkdir()
        shutil.copyfile(self.settings, second_root / "workflow-settings.toml")
        LedgerStore.create(second_root / "ledger" / "Skills主台账.xlsx")

        def packet_and_decision(root, candidate_id):
            source = root / f"{candidate_id}-source"
            source.mkdir()
            (source / "SKILL.md").write_text(candidate_id, encoding="utf-8")
            version, evidence = "e" * 40, (f"https://evidence.example/{candidate_id}",)
            snapshot = build_snapshot(SnapshotCandidate(candidate_id, version, source, evidence), root / f"{candidate_id}-snapshot")
            packet = build_review_packet({"id": candidate_id, "canonical_source": f"https://github.com/example/{candidate_id}", "license": "MIT", "security_grade": "SA"}, snapshot)
            decision = ReviewDecision(ObservedFacts(version, True, True, "MIT", f"https://github.com/example/{candidate_id}", evidence, "否", (), "无", "不使用", "SA", "全部通过（未实测）"), ProjectJudgments("正式推荐", True, True, 4, (True,)), candidate_id=candidate_id)
            return packet, decision

        packet_a, decision_a = packet_and_decision(self.root, "packet-a")
        packet_b, decision_b = packet_and_decision(second_root, "packet-b")
        first = self.coordinator()
        prepared_a = first.prepare(replace(self.request, review_packets={"packet-a": packet_a}))
        first.finalize(prepared_a, first.apply_reviews(prepared_a, (decision_a,)))

        second = RunCoordinator(root=second_root, discover=lambda request, staging: (
            SourceRun("SkillHub", "partial"), SourceRun("ClawHub", "failed"), SourceRun("GitHub", "failed"), SourceRun("Hugging Face Spaces", "failed"),
        ), report_builder=lambda prepared, staging: (), office_verifier=lambda prepared, artifacts: True)
        request_b = RunRequest(settings_path=second_root / "workflow-settings.toml", catalog_loader=lambda: object(), review_packets={"packet-b": packet_b})
        prepared_b = second.prepare(request_b)
        self.assertEqual(second.apply_reviews(prepared_b, (decision_b,)).applied_count, 1)
        second.abandon(prepared_b)

    def test_existing_generation_tree_change_before_commit_is_rejected(self):
        output = self.root / "output" / "generations" / "old"
        output.mkdir(parents=True)
        (output / "report.txt").write_text("old", encoding="utf-8")
        coordinator = self.coordinator()
        prepared = coordinator.prepare(self.request)
        (output / "report.txt").write_text("external modification", encoding="utf-8")
        with self.assertRaises(CoordinatorError):
            coordinator.finalize(prepared, coordinator.apply_reviews(prepared, ()))

    def test_before_commit_failure_removes_owned_generation_and_allows_same_run_id_retry(self):
        request = replace(self.request, requested_run_id="run-retry")
        first = self.coordinator(fail_at="before_commit")
        prepared = first.prepare(request)
        with self.assertRaises(RuntimeError):
            first.finalize(prepared, first.apply_reviews(prepared, ()))
        self.assertFalse((self.root / "output" / "generations" / "run-retry").exists())
        retry = self.coordinator()
        second = retry.prepare(request)
        retry.abandon(second)

    def test_post_commit_lock_release_error_returns_success_and_ledger_authority_remains_valid(self):
        coordinator = self.coordinator()
        prepared = coordinator.prepare(self.request)
        with patch.object(prepared._coordinator_token and coordinator._states[prepared.run_id].lock, "release", side_effect=OSError("release failure")):
            summary = coordinator.finalize(prepared, coordinator.apply_reviews(prepared, ()))
        self.assertFalse(summary.blocked)
        self.assertEqual(LedgerStore.load(self.root / "ledger" / "Skills主台账.xlsx").rows("运行记录")[-1]["状态"], "成功")

    def test_generation_manifest_and_ledger_share_delivery_hash_excluding_manifest(self):
        coordinator = self.coordinator()
        prepared = coordinator.prepare(self.request)
        summary = coordinator.finalize(prepared, coordinator.apply_reviews(prepared, ()))
        manifest = json.loads((summary.output_generation / "generation-manifest.json").read_text(encoding="utf-8"))
        record = LedgerStore.load(self.root / "ledger" / "Skills主台账.xlsx").rows("运行记录")[-1]
        self.assertIn("delivery_sha256", manifest)
        self.assertIn(manifest["delivery_sha256"], record["摘要"])

    def test_staging_delivery_mutation_after_office_callback_is_rejected_before_commit(self):
        def build_delivery(prepared, staging):
            delivery = staging / "deliveries"
            delivery.mkdir()
            report = delivery / "report.txt"
            report.write_text("approved", encoding="utf-8")
            return (report,)

        coordinator = self.coordinator(report_builder=build_delivery)
        prepared = coordinator.prepare(self.request)
        original_copytree = shutil.copytree

        def copy_then_background_write(source, destination, *args, **kwargs):
            result = original_copytree(source, destination, *args, **kwargs)
            (prepared.staging_dir / "deliveries" / "report.txt").write_text("late writer", encoding="utf-8")
            return result

        with patch("skill_maintainer.runner.shutil.copytree", side_effect=copy_then_background_write):
            with self.assertRaises(CoordinatorError):
                coordinator.finalize(prepared, coordinator.apply_reviews(prepared, ()))
        self.assertFalse((self.root / "output" / "generations" / prepared.run_id).exists())

    def test_prepare_rejects_missing_generation_referenced_by_last_successful_ledger_record(self):
        first = self.coordinator()
        prepared = first.prepare(self.request)
        summary = first.finalize(prepared, first.apply_reviews(prepared, ()))
        shutil.rmtree(summary.output_generation)
        with self.assertRaises(CoordinatorError):
            self.coordinator().prepare(self.request)

    def test_prepare_system_exit_cleans_staging_and_releases_project_writer_lock(self):
        crashing = self.coordinator()
        crashing.discover = lambda request, staging: (_ for _ in ()).throw(SystemExit("discovery crash"))
        request = replace(self.request, requested_run_id="run-prepare-system-exit")
        with self.assertRaises(SystemExit):
            crashing.prepare(request)
        self.assertFalse((self.root / ".runtime" / "staging" / "run-prepare-system-exit").exists())
        successor = self.coordinator()
        next_run = successor.prepare(self.request)
        successor.abandon(next_run)

    def test_new_formal_review_rebinds_three_platform_aliases_to_approved_stable_id(self):
        source = self.root / "alias-candidate"
        source.mkdir()
        (source / "SKILL.md").write_text("alias evidence", encoding="utf-8")
        version, evidence = "f" * 40, ("https://evidence.example/alias",)
        snapshot = build_snapshot(SnapshotCandidate("native-review", version, source, evidence), self.root / "alias-snapshot")
        packet = build_review_packet({"id": "native-review", "canonical_source": "https://github.com/example/alias", "license": "MIT", "security_grade": "SA"}, snapshot)
        approved = self.formal_row("APPROVED-ALIAS-1", version, snapshot.fixed_content_hash)
        approved.update({"Canonical source": "https://github.com/example/alias", "发现地址": "https://github.com/example/alias", "上游项目地址": "https://github.com/example/alias", "验证证据位置": "；".join(evidence), "本地专业软件或运行时依赖": "无", "本地脚本/插件接口": "不使用", "质量评分": 2})
        decision = ReviewDecision(
            ObservedFacts(version, True, True, "MIT", "https://github.com/example/alias", evidence, "否", (), "无", "不使用", "SA", "全部通过（未实测）"),
            ProjectJudgments("正式推荐", True, True, 4, (True,)), DerivedFields(quality_score=2, ledger_row=approved), "native-review",
        )
        evidence_file = self.root / "alias-discovery.json"
        evidence_file.write_text("{}", encoding="utf-8")
        candidates = tuple({"id": f"native-{platform}", "native_id": f"native-{platform}", "platform": platform, "source_url": f"https://{platform.lower().replace(' ', '-')}.example/alias", "canonical_source": "https://github.com/example/alias", "upstream_identity": "example/alias", "entry_path": "SKILL.md", "content_hash": snapshot.fixed_content_hash} for platform in ("SkillHub", "ClawHub", "GitHub"))
        batches = tuple(SourceRun(item["platform"], "complete", candidates=(item,), evidence_files=(evidence_file,)) for item in candidates) + (SourceRun("Hugging Face Spaces", "failed"),)
        coordinator = self.coordinator(discover=lambda request, staging: batches)
        prepared = coordinator.prepare(replace(self.request, review_packets={"native-review": packet}))
        coordinator.finalize(prepared, coordinator.apply_reviews(prepared, (decision,)))
        aliases = LedgerStore.load(self.root / "ledger" / "Skills主台账.xlsx").rows("来源别名")
        self.assertEqual({row["来源平台"] for row in aliases}, {"SkillHub", "ClawHub", "GitHub"})
        self.assertEqual({row["内部标识"] for row in aliases}, {"APPROVED-ALIAS-1"})

    def test_prepare_rejects_path_traversal_in_successful_generation_authority(self):
        first = self.coordinator()
        prepared = first.prepare(self.request)
        summary = first.finalize(prepared, first.apply_reviews(prepared, ()))
        outside = self.root / "outside-generation"
        shutil.copytree(summary.output_generation, outside)
        ledger_path = self.root / "ledger" / "Skills主台账.xlsx"
        ledger = LedgerStore.load(ledger_path)
        worksheet, columns = ledger.workbook["运行记录"], ledger._resolve_columns("运行记录")
        ledger._set_cell(worksheet.cell(worksheet.max_row, columns["摘要"]), "generation=output/generations/../../outside-generation;delivery_sha256=" + json.loads((outside / "generation-manifest.json").read_text(encoding="utf-8"))["delivery_sha256"] + ";manifest_sha256=" + sha256((outside / "generation-manifest.json").read_bytes()).hexdigest(), "摘要")
        ledger.save_staged(self.root / "ledger" / "path-injection.xlsx")
        ledger.workbook.close()
        shutil.copyfile(self.root / "ledger" / "path-injection.xlsx", ledger_path)
        with self.assertRaises(CoordinatorError):
            self.coordinator().prepare(self.request)

    def test_prepare_reclaims_only_valid_unreferenced_owned_orphan_generation(self):
        generation = self.root / "output" / "generations" / "run-orphan"
        generation.mkdir(parents=True)
        (generation / "report.txt").write_text("orphan", encoding="utf-8")
        helper = self.coordinator()
        digest = helper._tree_digest(generation)
        (generation / "generation-manifest.json").write_text(json.dumps({"run_id": "run-orphan", "delivery_sha256": digest,
            "files": helper._authority_files(generation)}), encoding="utf-8")
        coordinator = self.coordinator()
        prepared = coordinator.prepare(replace(self.request, requested_run_id="run-orphan"))
        self.assertFalse(generation.exists())
        coordinator.abandon(prepared)

    def test_generation_authority_files_are_pinned_during_ledger_commit(self):
        report_path: Path | None = None

        def build_delivery(prepared, staging):
            nonlocal report_path
            delivery = staging / "deliveries"
            delivery.mkdir()
            report_path = delivery / "report.txt"
            report_path.write_text("sealed", encoding="utf-8")
            return (report_path,)

        coordinator = self.coordinator(report_builder=build_delivery)
        prepared = coordinator.prepare(self.request)
        original_replace = os.replace

        def replace_with_late_write(source, destination):
            if Path(destination) == self.root / "ledger" / "Skills主台账.xlsx":
                generation_report = self.root / "output" / "generations" / prepared.run_id / "report.txt"
                with self.assertRaises(PermissionError):
                    generation_report.write_text("late mutation", encoding="utf-8")
            return original_replace(source, destination)

        with patch("skill_maintainer.runner.os.replace", side_effect=replace_with_late_write):
            coordinator.finalize(prepared, coordinator.apply_reviews(prepared, ()))

    def test_finalize_catalog_system_exit_cleans_state_staging_and_writer_lock(self):
        calls = 0

        def catalog():
            nonlocal calls
            calls += 1
            if calls > 2:
                raise SystemExit("catalog final preflight crash")
            return {"revision": "prepared"}

        request = replace(self.request, catalog_loader=catalog, requested_run_id="run-final-system-exit")
        coordinator = self.coordinator()
        prepared = coordinator.prepare(request)
        reviews = coordinator.apply_reviews(prepared, ())
        with self.assertRaises(SystemExit):
            coordinator.finalize(prepared, reviews)
        self.assertNotIn(prepared.run_id, coordinator._states)
        self.assertFalse(prepared.staging_dir.exists())
        successor = self.coordinator()
        next_run = successor.prepare(self.request)
        successor.abandon(next_run)

    def test_callback_junction_is_rejected_before_office_verifier_runs(self):
        outside = self.root / "outside-callback"
        outside.mkdir()
        (outside / "report.txt").write_text("external", encoding="utf-8")
        verifier_calls: list[tuple[Path, ...]] = []

        def build_delivery(prepared, staging):
            delivery = staging / "deliveries"
            delivery.mkdir()
            junction = delivery / "junction"
            try:
                junction.symlink_to(outside, target_is_directory=True)
            except OSError as error:
                self.skipTest(f"test filesystem cannot create a link: {error}")
            return (junction / "report.txt",)

        coordinator = self.coordinator(report_builder=build_delivery, office_verifier=lambda prepared, artifacts: verifier_calls.append(artifacts) or True)
        prepared = coordinator.prepare(self.request)
        with self.assertRaises(CoordinatorError):
            coordinator.finalize(prepared, coordinator.apply_reviews(prepared, ()))
        self.assertEqual(verifier_calls, [])

    def test_system_exit_after_ledger_replace_preserves_committed_generation_and_releases_lock(self):
        coordinator = self.coordinator()
        prepared = coordinator.prepare(self.request)
        original_replace = os.replace

        def replace_then_crash(source, destination):
            result = original_replace(source, destination)
            if Path(destination) == self.root / "ledger" / "Skills主台账.xlsx":
                raise SystemExit("after durable ledger commit")
            return result

        with patch("skill_maintainer.runner.os.replace", side_effect=replace_then_crash):
            with self.assertRaises(SystemExit):
                coordinator.finalize(prepared, coordinator.apply_reviews(prepared, ()))
        generation = self.root / "output" / "generations" / prepared.run_id
        self.assertTrue(generation.is_dir())
        records = LedgerStore.load(self.root / "ledger" / "Skills主台账.xlsx").rows("运行记录")
        self.assertEqual(records[-1]["状态"], "成功")
        successor = self.coordinator()
        next_run = successor.prepare(self.request)
        successor.abandon(next_run)


if __name__ == "__main__":
    unittest.main()
