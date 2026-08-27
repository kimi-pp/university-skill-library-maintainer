import tempfile
import unittest
import zipfile
from datetime import date, datetime
from hashlib import sha256
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.table import Table

from skill_maintainer.catalog import TaskProfileReadError, load_task_profiles_from_ledger
from skill_maintainer.ledger import LedgerStore, ProfessionalTaskMapUpgradeError, upgrade_professional_task_maps
from skill_maintainer.ledger_schema import (
    CURRENT_SKILL_COLUMNS,
    ERROR_DUPLICATE_CANONICAL_SOURCE,
    ERROR_DUPLICATE_STABLE_ID,
    ERROR_EXTRA_WORKSHEET,
    ERROR_FORMAL_INVALID_QUALITY_SCORE,
    ERROR_FORMAL_INVALID_SECURITY_GRADE,
    ERROR_FORMAL_INVALID_VALIDATION_STATUS,
    ERROR_FORMAL_MISSING_REQUIRED_FACT,
    ERROR_INVALID_REMOTE_CALL_FLAG,
    ERROR_INVALID_SOURCE_PLATFORM,
    ERROR_NON_FORMAL_CURRENT_SKILL,
    SHEET_SPECS_BY_NAME,
    ERROR_FORMAL_UNKNOWN_LICENSE,
    ERROR_LOCAL_SOFTWARE_IN_REMOTE_ENDPOINT,
    ERROR_MISSING_FIXED_VERSION,
    ERROR_REMOTE_ENDPOINT_REQUIRED,
)


EXPECTED_SHEETS = (
    "当前Skill", "来源别名", "专业任务映射", "版本历史", "候选观察",
    "目录基线", "来源水位", "运行记录", "字段说明",
)

EXPECTED_CURRENT_SKILL_COLUMNS = (
    "内部标识", "Skill名称", "规范名称", "入库层级", "功能一级分类", "功能二级标签", "关联分类", "原生生态", "来源形态", "来源平台", "发现地址", "收集日期", "Canonical source", "上游项目地址", "Skill入口路径", "发布者", "固定版本", "固定版本内容指纹", "许可证", "简要功能", "详细功能摘要", "适用用户角色", "典型高校场景", "Codex兼容等级", "适配建议", "关联资源类型", "关联资源地址", "外部依赖", "外部联网/API 调用", "远程服务端点", "本地专业软件或运行时依赖", "本地脚本/插件接口", "可执行行为", "网络与数据行为", "凭据行为", "文件行为", "安全等级", "安全限制条件", "最近更新", "维护状态", "风险提示", "替代方案", "验证级别", "验证状态", "验证证据位置", "最近核验日期", "推荐优先级", "接入难度", "实施准备度", "质量评分", "重复或关联条目", "备注",
)

EXPECTED_SOURCE_ALIAS_COLUMNS = (
    "别名标识", "内部标识", "来源平台", "来源地址", "Canonical source", "关系类型", "去重依据", "记录日期",
)

EXPECTED_PROFESSIONAL_TASK_MAP_COLUMNS = (
    "映射标识", "内部标识", "专业代码", "专业名称", "专业任务", "输入", "输出", "适用理由", "使用限制", "相关度",
    "专业别名", "核心课程", "研究方法", "工作任务", "成果或数据对象", "软件/数据库/流程",
)


def formal_row(number: int = 1, **overrides):
    row = {
        "内部标识": f"GH-01-{number:04d}",
        "Skill名称": f"skill-{number}",
        "规范名称": f"技能 {number}",
        "入库层级": "正式",
        "功能一级分类": "01",
        "功能二级标签": "学术写作",
        "关联分类": "",
        "原生生态": "Codex",
        "来源形态": "Skill",
        "来源平台": "GitHub",
        "发现地址": f"https://example.edu/discovery/{number}",
        "收集日期": date(2026, 8, 20),
        "Canonical source": f"https://example.edu/skills/{number}",
        "上游项目地址": f"https://example.edu/upstream/{number}",
        "Skill入口路径": "SKILL.md",
        "发布者": "示例维护者",
        "固定版本": f"v{number}.0.0",
        "固定版本内容指纹": f"sha256:{number:064x}",
        "许可证": "MIT",
        "简要功能": "生成结构化教学材料。",
        "详细功能摘要": "根据已核验的输入生成材料，并保留人工复核边界。",
        "适用用户角色": "教学人员",
        "典型高校场景": "课程材料准备",
        "Codex兼容等级": "A",
        "适配建议": "直接使用",
        "关联资源类型": "",
        "关联资源地址": "",
        "外部依赖": "未见明确外部依赖",
        "外部联网/API 调用": "否",
        "远程服务端点": "",
        "本地专业软件或运行时依赖": "无",
        "本地脚本/插件接口": "不使用",
        "可执行行为": "未见可执行载荷",
        "网络与数据行为": "未见网络请求",
        "凭据行为": "未见凭据读取",
        "文件行为": "未见文件写入",
        "安全等级": "SA",
        "安全限制条件": "无额外限制",
        "最近更新": date(2026, 8, 1),
        "维护状态": "活跃",
        "风险提示": "未见明显风险",
        "替代方案": "无",
        "验证级别": "二级包内容验证",
        "最近核验日期": date(2026, 8, 27),
        "验证状态": "全部通过（未实测）",
        "验证证据位置": f"https://example.edu/evidence/{number}",
        "推荐优先级": "高",
        "接入难度": "A",
        "实施准备度": "可直接使用",
        "质量评分": 5,
        "重复或关联条目": "",
        "备注": "",
    }
    row.update(overrides)
    return row


def create_legacy_task_map_workbook(path: Path) -> None:
    """构造一个具备命名表的实际 10 列旧版主台账。"""
    store = LedgerStore.create(path)
    store.append_rows("专业任务映射", [{
        "映射标识": "PROFILE-0818",
        "内部标识": "LEGACY-0818",
        "专业代码": "0818",
        "专业名称": "交通运输类",
        "专业任务": "交通运行分析",
        "输入": "OD矩阵",
        "输出": "分析报告",
        "适用理由": "旧版映射",
        "使用限制": "需人工复核",
        "相关度": 5,
    }])
    worksheet = store.workbook["专业任务映射"]
    worksheet.delete_cols(11, 6)
    table = next(iter(worksheet.tables.values()))
    table.ref = "A1:J2"
    table.autoFilter.ref = table.ref
    store.workbook.save(path)
    store.workbook.close()


class LedgerStoreTest(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        self.addCleanup(self.tempdir.cleanup)
        self.production_path = Path(self.tempdir.name) / "production.xlsx"
        self.store = LedgerStore.create(self.production_path)

    def test_creates_exact_named_sheets_and_named_column_tables(self):
        self.assertEqual(tuple(self.store.workbook.sheetnames), EXPECTED_SHEETS)
        self.assertEqual(CURRENT_SKILL_COLUMNS, EXPECTED_CURRENT_SKILL_COLUMNS)
        self.assertEqual(SHEET_SPECS_BY_NAME["来源别名"].columns, EXPECTED_SOURCE_ALIAS_COLUMNS)
        self.assertEqual(SHEET_SPECS_BY_NAME["专业任务映射"].columns, EXPECTED_PROFESSIONAL_TASK_MAP_COLUMNS)
        for sheet_name in EXPECTED_SHEETS:
            with self.subTest(sheet=sheet_name):
                worksheet = self.store.workbook[sheet_name]
                self.assertEqual(worksheet.freeze_panes, "A2")
                self.assertEqual(len(worksheet.tables), 1)
                self.assertTrue(next(iter(worksheet.tables.values())).ref.endswith("2"))
                self.assertEqual(self.store.rows(sheet_name), [])

    def test_first_append_reuses_excel_compatible_empty_table_placeholder(self):
        self.store.append_rows("当前Skill", [formal_row(1)])
        worksheet = self.store.workbook["当前Skill"]
        table = next(iter(worksheet.tables.values()))

        self.assertEqual(len(self.store.rows("当前Skill")), 1)
        self.assertEqual(worksheet.cell(2, 1).value, "GH-01-0001")
        self.assertTrue(table.ref.endswith("2"))

    def test_520_formal_rows_survive_saved_reopen_with_table_filter_hyperlink_and_date(self):
        self.store.append_rows("当前Skill", [formal_row(number) for number in range(1, 521)])
        staging_path = Path(self.tempdir.name) / "520-fixture.xlsx"
        checksum = self.store.save_staged(staging_path)

        reopened = LedgerStore.load(staging_path)
        worksheet = reopened.workbook["当前Skill"]
        table = next(iter(worksheet.tables.values()))
        columns = {cell.value: cell.column for cell in worksheet[1]}

        self.assertEqual(len(reopened.rows("当前Skill")), 520)
        self.assertEqual(len({row["内部标识"] for row in reopened.rows("当前Skill")}), 520)
        self.assertTrue(table.ref.endswith("521"))
        self.assertIsNotNone(table.autoFilter)
        self.assertEqual(table.autoFilter.ref, table.ref)
        self.assertIsNone(worksheet.auto_filter.ref)
        self.assertEqual(worksheet.freeze_panes, "A2")
        self.assertEqual(
            worksheet.cell(2, columns["Canonical source"]).hyperlink.target,
            "https://example.edu/skills/1",
        )
        self.assertIsInstance(worksheet.cell(2, columns["最近核验日期"]).value, date)
        self.assertEqual(len(checksum), 64)
        self.assertEqual(reopened.validate(), [])

        with zipfile.ZipFile(staging_path) as archive:
            worksheet_xml = archive.read("xl/worksheets/sheet1.xml")
            table_xml = archive.read("xl/tables/table1.xml")
        self.assertNotIn(b"<autoFilter", worksheet_xml)
        self.assertIn(b'<autoFilter ref="A1:AZ521"', table_xml)

    def test_520_profile_rows_survive_saved_reopen_with_named_columns_and_ooxml_table_filter(self):
        self.store.append_rows("专业任务映射", [{
            "映射标识": f"PROFILE-{number:04d}",
            "专业代码": f"P{number:04d}",
            "专业名称": f"专业类 {number}",
            "专业别名": f"别名 {number}",
            "核心课程": "核心课程",
            "研究方法": "研究方法",
            "工作任务": "工作任务",
            "成果或数据对象": "数据对象",
            "软件/数据库/流程": "分析流程",
        } for number in range(1, 521)])
        staging_path = Path(self.tempdir.name) / "520-profiles.xlsx"
        self.store.save_staged(staging_path)

        reopened = LedgerStore.load(staging_path)
        worksheet = reopened.workbook["专业任务映射"]
        table = next(iter(worksheet.tables.values()))
        columns = {cell.value: cell.column for cell in worksheet[1]}

        self.assertEqual(len(reopened.rows("专业任务映射")), 520)
        self.assertEqual(worksheet.cell(2, columns["专业代码"]).value, "P0001")
        self.assertEqual(worksheet.freeze_panes, "A2")
        self.assertEqual(table.autoFilter.ref, table.ref)
        self.assertTrue(table.ref.endswith("521"))
        with zipfile.ZipFile(staging_path) as archive:
            worksheet_xml = archive.read("xl/worksheets/sheet3.xml")
            table_xml = archive.read("xl/tables/table3.xml")
        self.assertNotIn(b"<autoFilter", worksheet_xml)
        self.assertIn(b'<autoFilter ref="A1:P521"', table_xml)

    def test_legacy_task_map_upgrade_is_staged_immutable_and_requires_profile_completion(self):
        legacy_path = Path(self.tempdir.name) / "legacy-10-columns.xlsx"
        staged_path = Path(self.tempdir.name) / "upgraded-16-columns.xlsx"
        current_copy_path = Path(self.tempdir.name) / "current-copy.xlsx"
        create_legacy_task_map_workbook(legacy_path)
        source_sha = sha256(legacy_path.read_bytes()).hexdigest()

        summary = upgrade_professional_task_maps(legacy_path, staged_path)

        self.assertTrue(summary.upgraded_legacy_schema)
        self.assertEqual(summary.source_sha256, source_sha)
        self.assertEqual(summary.legacy_profile_rows_requiring_completion, 1)
        self.assertEqual(sha256(legacy_path.read_bytes()).hexdigest(), source_sha)
        self.assertEqual(LedgerStore.load(legacy_path).workbook["专业任务映射"].max_column, 10)

        upgraded = LedgerStore.load(staged_path)
        self.assertEqual(upgraded.workbook["专业任务映射"].max_column, 16)
        self.assertEqual(upgraded.rows("专业任务映射")[0]["软件/数据库/流程"], "需补录")
        self.assertEqual(upgraded.validate(), [])
        upgraded.workbook.close()
        with self.assertRaises(TaskProfileReadError):
            load_task_profiles_from_ledger(staged_path)

        copied_summary = upgrade_professional_task_maps(staged_path, current_copy_path)
        self.assertFalse(copied_summary.upgraded_legacy_schema)
        self.assertEqual(copied_summary.legacy_profile_rows_requiring_completion, 0)
        self.assertEqual(copied_summary.source_sha256, copied_summary.staging_sha256)
        self.assertEqual(LedgerStore.load(current_copy_path).workbook["专业任务映射"].max_column, 16)

    def test_legacy_upgrade_keeps_requested_output_absent_or_unmodified_when_source_is_invalid(self):
        invalid_source = Path(self.tempdir.name) / "invalid-legacy.xlsx"
        output = Path(self.tempdir.name) / "must-not-change.xlsx"
        create_legacy_task_map_workbook(invalid_source)
        workbook = LedgerStore.load(invalid_source).workbook
        workbook["专业任务映射"]["A1"] = "未知字段"
        workbook.save(invalid_source)
        workbook.close()
        output.write_bytes(b"unchanged target bytes")

        with self.assertRaises(ProfessionalTaskMapUpgradeError):
            upgrade_professional_task_maps(invalid_source, output)

        self.assertEqual(output.read_bytes(), b"unchanged target bytes")

    def test_validation_reports_stable_codes_for_formal_row_failures(self):
        cases = (
            (
                [formal_row(1), formal_row(1, **{"Canonical source": "https://example.edu/other"})],
                ERROR_DUPLICATE_STABLE_ID,
            ),
            (
                [formal_row(1), formal_row(2, **{"Canonical source": "https://example.edu/skills/1"})],
                ERROR_DUPLICATE_CANONICAL_SOURCE,
            ),
            ([formal_row(1, **{"固定版本": ""})], ERROR_MISSING_FIXED_VERSION),
            ([formal_row(1, **{"许可证": "未明确"})], ERROR_FORMAL_UNKNOWN_LICENSE),
            (
                [formal_row(1, **{"外部联网/API 调用": "是", "远程服务端点": ""})],
                ERROR_REMOTE_ENDPOINT_REQUIRED,
            ),
            (
                [formal_row(1, **{"外部联网/API 调用": "是", "远程服务端点": "Abaqus"})],
                ERROR_LOCAL_SOFTWARE_IN_REMOTE_ENDPOINT,
            ),
        )
        for rows, error_code in cases:
            with self.subTest(error_code=error_code):
                store = LedgerStore.create(Path(self.tempdir.name) / f"{error_code}.xlsx")
                store.append_rows("当前Skill", rows)
                self.assertIn(error_code, store.validate())

    def test_current_skill_rejects_nonformal_rows_and_missing_or_invalid_formal_facts(self):
        cases = (
            ({"内部标识": "GH-01-0001", "Skill名称": "minimal", "入库层级": "正式"}, ERROR_FORMAL_MISSING_REQUIRED_FACT),
            (formal_row(1, **{"入库层级": "观察"}), ERROR_NON_FORMAL_CURRENT_SKILL),
            (formal_row(1, **{"验证状态": "前两步通过"}), ERROR_FORMAL_INVALID_VALIDATION_STATUS),
            (formal_row(1, **{"安全等级": "SB-A"}), ERROR_FORMAL_INVALID_SECURITY_GRADE),
            (formal_row(1, **{"质量评分": 1}), ERROR_FORMAL_INVALID_QUALITY_SCORE),
            (formal_row(1, **{"质量评分": 2.5}), ERROR_FORMAL_INVALID_QUALITY_SCORE),
            (formal_row(1, **{"来源平台": "UnknownMarketplace"}), ERROR_INVALID_SOURCE_PLATFORM),
            (formal_row(1, **{"外部联网/API 调用": "可能"}), ERROR_INVALID_REMOTE_CALL_FLAG),
        )
        for row, error_code in cases:
            with self.subTest(error_code=error_code):
                store = LedgerStore.create(Path(self.tempdir.name) / f"formal-{error_code}.xlsx")
                store.append_rows("当前Skill", [row])
                self.assertIn(error_code, store.validate())

    def test_formal_license_placeholders_are_rejected_but_authorized_tested_status_is_kept(self):
        placeholders = ("待确认", "无许可证声明", "未明确", "未知", " unknown ", "N/A", "NONE", "null", "-")
        for license_name in placeholders:
            with self.subTest(license_name=license_name):
                store = LedgerStore.create(Path(self.tempdir.name) / f"license-{len(license_name)}.xlsx")
                store.append_rows("当前Skill", [formal_row(1, **{"许可证": license_name})])
                self.assertIn(ERROR_FORMAL_UNKNOWN_LICENSE, store.validate())
        self.store.append_rows("当前Skill", [formal_row(1, **{"验证状态": "全部通过（已实测）"})])
        self.assertNotIn(ERROR_FORMAL_INVALID_VALIDATION_STATUS, self.store.validate())

    def test_formal_license_must_be_a_nonempty_plain_string(self):
        self.store.append_rows("当前Skill", [formal_row(1, **{"许可证": 123})])
        self.assertIn(ERROR_FORMAL_UNKNOWN_LICENSE, self.store.validate())

    def test_validation_rejects_extra_sheets_and_invalid_table_structure_before_staged_save(self):
        self.store.workbook.create_sheet("多余工作表")
        worksheet = self.store.workbook["当前Skill"]
        worksheet.add_table(Table(displayName="UnexpectedTable", ref="A1:AW2"))
        worksheet.auto_filter.ref = "A1:AW2"
        errors = self.store.validate()

        self.assertIn(ERROR_EXTRA_WORKSHEET, errors)
        self.assertIn("台账错误-命名表数量错误-当前Skill", errors)
        self.assertIn("台账错误-工作表筛选不允许-当前Skill", errors)
        with self.assertRaises(ValueError):
            self.store.save_staged(Path(self.tempdir.name) / "broken-structure.xlsx")

    def test_run_record_datetimes_keep_seconds_display_format(self):
        self.store.append_rows("运行记录", [{
            "运行标识": "run-001",
            "运行类型": "维护",
            "开始时间": datetime(2026, 8, 27, 22, 0, 1),
            "成功完成时间": datetime(2026, 8, 27, 22, 1, 2),
            "状态": "成功",
            "摘要": "完成",
            "快照SHA-256": "a" * 64,
        }])
        staging_path = Path(self.tempdir.name) / "run-record-datetimes.xlsx"
        self.store.save_staged(staging_path)
        worksheet = LedgerStore.load(staging_path).workbook["运行记录"]
        columns = {cell.value: cell.column for cell in worksheet[1]}
        self.assertEqual(worksheet.cell(2, columns["开始时间"]).number_format, "yyyy-mm-dd hh:mm:ss")
        self.assertEqual(worksheet.cell(2, columns["成功完成时间"]).number_format, "yyyy-mm-dd hh:mm:ss")

    def test_rejects_duplicate_and_unknown_headers_when_resolving_named_columns(self):
        worksheet = self.store.workbook["当前Skill"]
        worksheet.cell(1, 2).value = "内部标识"
        self.assertIn("台账错误-重复表头-当前Skill", self.store.validate())

        worksheet.cell(1, 2).value = "未知字段"
        self.assertIn("台账错误-未知表头-当前Skill", self.store.validate())

    def test_upsert_updates_by_stable_id_and_snapshot_uses_named_rows(self):
        self.store.upsert_skill(formal_row(1))
        self.store.upsert_skill(formal_row(1, **{"规范名称": "更新后的技能"}))
        snapshot = self.store.current_snapshot()

        self.assertEqual(len(self.store.rows("当前Skill")), 1)
        self.assertEqual(self.store.rows("当前Skill")[0]["规范名称"], "更新后的技能")
        self.assertEqual(snapshot.row_counts["当前Skill"], 1)
        self.assertIn("当前Skill", snapshot.sheet_names)

    def test_staged_save_never_overwrites_the_loaded_production_ledger(self):
        loaded = LedgerStore.load(self.production_path)
        with self.assertRaises(ValueError):
            loaded.save_staged(self.production_path)


if __name__ == "__main__":
    unittest.main()
