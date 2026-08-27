"""基于 openpyxl 的 Excel 主台账读写与校验。"""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from hashlib import sha256
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter, range_boundaries

from .ledger_schema import (
    ALLOWED_SOURCE_PLATFORMS,
    ERROR_DUPLICATE_CANONICAL_SOURCE,
    ERROR_DUPLICATE_STABLE_ID,
    ERROR_EXTRA_WORKSHEET,
    ERROR_FORMAL_INVALID_QUALITY_SCORE,
    ERROR_FORMAL_INVALID_SECURITY_GRADE,
    ERROR_FORMAL_INVALID_VALIDATION_STATUS,
    ERROR_FORMAL_MISSING_REQUIRED_FACT,
    ERROR_FORMAL_UNKNOWN_LICENSE,
    ERROR_INVALID_REMOTE_CALL_FLAG,
    ERROR_INVALID_SOURCE_PLATFORM,
    ERROR_LOCAL_SOFTWARE_IN_REMOTE_ENDPOINT,
    ERROR_MISSING_FIXED_VERSION,
    ERROR_REMOTE_ENDPOINT_REQUIRED,
    ERROR_NON_FORMAL_CURRENT_SKILL,
    CURRENT_SKILL_REQUIRED_COLUMNS,
    SHEET_SPECS,
    SHEET_SPECS_BY_NAME,
    SheetSpec,
)


@dataclass(frozen=True)
class LedgerSnapshot:
    path: Path | None
    sha256: str | None
    row_counts: dict[str, int]
    sheet_names: tuple[str, ...]


def validate_current_skill_row(row: Mapping[str, Any]) -> list[str]:
    """返回单条当前 Skill 正式语义校验的稳定错误码。"""
    errors: list[str] = []
    if str(row["入库层级"] or "").strip() != "正式":
        return [ERROR_NON_FORMAL_CURRENT_SKILL]
    if any(not LedgerStore._has_value(row[column]) for column in CURRENT_SKILL_REQUIRED_COLUMNS):
        errors.append(ERROR_FORMAL_MISSING_REQUIRED_FACT)
    if not str(row["固定版本"] or "").strip():
        errors.append(ERROR_MISSING_FIXED_VERSION)
    license_value = row["许可证"]
    license_name = license_value.strip().casefold() if type(license_value) is str else ""
    if type(license_value) is not str or license_name in {"", "待确认", "无许可证声明", "未明确", "未知", "unknown", "n/a", "none", "null", "-"}:
        errors.append(ERROR_FORMAL_UNKNOWN_LICENSE)
    if row["来源平台"] not in ALLOWED_SOURCE_PLATFORMS:
        errors.append(ERROR_INVALID_SOURCE_PLATFORM)
    remote_call = str(row["外部联网/API 调用"] or "").strip()
    endpoint = str(row["远程服务端点"] or "").strip()
    if remote_call not in {"是", "否"}:
        errors.append(ERROR_INVALID_REMOTE_CALL_FLAG)
    if remote_call == "是" and not endpoint:
        errors.append(ERROR_REMOTE_ENDPOINT_REQUIRED)
    endpoint_lower = endpoint.lower()
    if remote_call == "是" and any(name in endpoint_lower for name in ("abaqus", "ansys", "matlab", "autocad")):
        errors.append(ERROR_LOCAL_SOFTWARE_IN_REMOTE_ENDPOINT)
    if row["验证状态"] not in {"全部通过（未实测）", "全部通过（已实测）"}:
        errors.append(ERROR_FORMAL_INVALID_VALIDATION_STATUS)
    quality_score = row["质量评分"]
    if row["安全等级"] not in {"SA", "SB"}:
        errors.append(ERROR_FORMAL_INVALID_SECURITY_GRADE)
    if isinstance(quality_score, bool) or not isinstance(quality_score, int) or quality_score < 2 or quality_score > 5:
        errors.append(ERROR_FORMAL_INVALID_QUALITY_SCORE)
    return list(dict.fromkeys(errors))


class LedgerStore:
    """所有字段均通过表头名称读取，避免依赖固定列号。"""

    def __init__(self, workbook: Workbook, source_path: Path | None = None):
        self.workbook = workbook
        self.source_path = source_path.resolve() if source_path else None
        self._last_staged_path: Path | None = None
        self._last_staged_sha256: str | None = None

    @classmethod
    def create(cls, path: str | Path) -> "LedgerStore":
        target = Path(path).resolve()
        if target.suffix.lower() != ".xlsx":
            raise ValueError("主台账必须使用 .xlsx 文件")
        target.parent.mkdir(parents=True, exist_ok=True)
        workbook = Workbook()
        workbook.remove(workbook.active)
        for spec in SHEET_SPECS:
            worksheet = workbook.create_sheet(spec.name)
            cls._initialize_sheet(worksheet, spec)
        workbook.save(target)
        return cls(workbook, target)

    @classmethod
    def load(cls, path: str | Path) -> "LedgerStore":
        target = Path(path).resolve()
        return cls(load_workbook(target, data_only=False), target)

    @staticmethod
    def _initialize_sheet(worksheet, spec: SheetSpec) -> None:
        for index, column_name in enumerate(spec.columns, start=1):
            cell = worksheet.cell(1, index, column_name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            worksheet.column_dimensions[get_column_letter(index)].width = max(14, min(34, len(column_name) * 2 + 4))
        worksheet.freeze_panes = "A2"
        # Excel 会修复只含表头的 Table；保留一个全空数据行而不把它当业务记录。
        for index in range(1, len(spec.columns) + 1):
            worksheet.cell(2, index)
        table = Table(displayName=spec.table_name, ref=f"A1:{get_column_letter(len(spec.columns))}2")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        worksheet.add_table(table)

    def _spec(self, sheet: str) -> SheetSpec:
        try:
            return SHEET_SPECS_BY_NAME[sheet]
        except KeyError as error:
            raise KeyError(f"未知工作表：{sheet}") from error

    def _resolve_columns(self, sheet: str, raise_on_error: bool = True) -> dict[str, int]:
        spec = self._spec(sheet)
        worksheet = self.workbook[sheet]
        headers = [worksheet.cell(1, column).value for column in range(1, worksheet.max_column + 1)]
        errors: list[str] = []
        known = set(spec.columns)
        seen: set[str] = set()
        for header in headers:
            if header in seen:
                errors.append(f"台账错误-重复表头-{sheet}")
            seen.add(header)
            if header not in known:
                errors.append(f"台账错误-未知表头-{sheet}")
        if set(headers) != known:
            errors.append(f"台账错误-表头不完整-{sheet}")
        if errors and raise_on_error:
            raise ValueError("；".join(dict.fromkeys(errors)))
        return {header: index for index, header in enumerate(headers, start=1) if header in known}

    def _table(self, sheet: str) -> Table | None:
        spec = self._spec(sheet)
        return self.workbook[sheet].tables.get(spec.table_name)

    def _resize_table(self, sheet: str) -> None:
        spec = self._spec(sheet)
        worksheet = self.workbook[sheet]
        table = self._table(sheet)
        if table is None:
            table = Table(displayName=spec.table_name, ref="A1:A1")
            worksheet.add_table(table)
        table.ref = f"A1:{get_column_letter(len(spec.columns))}{max(2, worksheet.max_row)}"
        table.autoFilter.ref = table.ref

    @staticmethod
    def _set_cell(cell, value: Any, column_name: str) -> None:
        cell.value = value
        cell.hyperlink = None
        if isinstance(value, str) and value.startswith(("https://", "http://")):
            cell.hyperlink = value
            cell.style = "Hyperlink"
        if isinstance(value, datetime):
            cell.number_format = "yyyy-mm-dd hh:mm:ss"
        elif isinstance(value, date):
            cell.number_format = "yyyy-mm-dd"
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    def append_rows(self, sheet: str, rows: Iterable[Mapping[str, Any]]) -> None:
        spec = self._spec(sheet)
        columns = self._resolve_columns(sheet)
        worksheet = self.workbook[sheet]
        for row in rows:
            unknown = set(row) - set(spec.columns)
            if unknown:
                raise KeyError(f"{sheet} 包含未知字段：{sorted(unknown)}")
            row_number = next(
                (
                    candidate
                    for candidate in range(2, worksheet.max_row + 1)
                    if all(worksheet.cell(candidate, index).value in (None, "") for index in columns.values())
                ),
                worksheet.max_row + 1,
            )
            for column_name in spec.columns:
                self._set_cell(worksheet.cell(row_number, columns[column_name]), row.get(column_name), column_name)
        self._resize_table(sheet)

    def rows(self, sheet: str) -> list[dict[str, Any]]:
        spec = self._spec(sheet)
        columns = self._resolve_columns(sheet)
        worksheet = self.workbook[sheet]
        result = []
        for row_number in range(2, worksheet.max_row + 1):
            values = {name: worksheet.cell(row_number, index).value for name, index in columns.items()}
            if any(value not in (None, "") for value in values.values()):
                result.append({name: values.get(name) for name in spec.columns})
        return result

    def upsert_skill(self, row: Mapping[str, Any]) -> None:
        stable_id = row.get("内部标识")
        if not stable_id:
            raise ValueError("当前Skill upsert 需要内部标识")
        spec = self._spec("当前Skill")
        columns = self._resolve_columns(spec.name)
        worksheet = self.workbook[spec.name]
        for row_number in range(2, worksheet.max_row + 1):
            if worksheet.cell(row_number, columns["内部标识"]).value == stable_id:
                unknown = set(row) - set(spec.columns)
                if unknown:
                    raise KeyError(f"当前Skill 包含未知字段：{sorted(unknown)}")
                for column_name in spec.columns:
                    if column_name in row:
                        self._set_cell(worksheet.cell(row_number, columns[column_name]), row[column_name], column_name)
                return
        self.append_rows(spec.name, [row])

    def validate(self) -> list[str]:
        errors: list[str] = []
        current_skill_rows: list[dict[str, Any]] | None = None
        expected_sheet_names = {spec.name for spec in SHEET_SPECS}
        if set(self.workbook.sheetnames) - expected_sheet_names:
            errors.append(ERROR_EXTRA_WORKSHEET)
        for spec in SHEET_SPECS:
            if spec.name not in self.workbook.sheetnames:
                errors.append(f"台账错误-缺少工作表-{spec.name}")
                continue
            worksheet = self.workbook[spec.name]
            try:
                columns = self._resolve_columns(spec.name)
            except ValueError as error:
                errors.extend(str(error).split("；"))
                continue
            table = self._table(spec.name)
            if len(worksheet.tables) != 1:
                errors.append(f"台账错误-命名表数量错误-{spec.name}")
            if table is None:
                errors.append(f"台账错误-缺少数据表-{spec.name}")
            else:
                expected_ref = f"A1:{get_column_letter(len(spec.columns))}{max(2, worksheet.max_row)}"
                if table.ref != expected_ref:
                    errors.append(f"台账错误-数据表范围不一致-{spec.name}")
                if table.autoFilter is None or table.autoFilter.ref != table.ref:
                    errors.append(f"台账错误-数据表筛选范围不一致-{spec.name}")
                if worksheet.auto_filter.ref is not None:
                    errors.append(f"台账错误-工作表筛选不允许-{spec.name}")
            for unique_key in spec.unique_keys:
                seen: set[tuple[str, ...]] = set()
                for row in self.rows(spec.name):
                    values = tuple(str(row[key]).strip() for key in unique_key)
                    if not all(values):
                        continue
                    if values in seen:
                        if spec.name == "当前Skill" and unique_key == ("内部标识",):
                            errors.append(ERROR_DUPLICATE_STABLE_ID)
                        elif spec.name == "当前Skill" and unique_key == ("Canonical source",):
                            errors.append(ERROR_DUPLICATE_CANONICAL_SOURCE)
                        else:
                            errors.append(f"台账错误-重复唯一键-{spec.name}")
                    seen.add(values)
            if spec.name == "当前Skill":
                current_skill_rows = self.rows(spec.name)

        if current_skill_rows is not None:
            for row in current_skill_rows:
                errors.extend(validate_current_skill_row(row))
        return list(dict.fromkeys(errors))

    @staticmethod
    def _has_value(value: Any) -> bool:
        return value is not None and (not isinstance(value, str) or bool(value.strip()))

    def save_staged(self, path: str | Path) -> str:
        target = Path(path).resolve()
        if target.suffix.lower() != ".xlsx":
            raise ValueError("暂存台账必须使用 .xlsx 文件")
        if self.source_path is not None and target == self.source_path:
            raise ValueError("暂存路径不得等于当前生产台账路径")
        target.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(target)
        self.workbook.close()
        verified = LedgerStore.load(target)
        errors = verified.validate()
        if errors:
            verified.workbook.close()
            raise ValueError("暂存台账校验失败：" + "；".join(errors))
        digest = sha256(target.read_bytes()).hexdigest()
        self.workbook = verified.workbook
        self._last_staged_path = target
        self._last_staged_sha256 = digest
        return digest

    def current_snapshot(self) -> LedgerSnapshot:
        return LedgerSnapshot(
            path=self._last_staged_path or self.source_path,
            sha256=self._last_staged_sha256,
            row_counts={spec.name: len(self.rows(spec.name)) for spec in SHEET_SPECS},
            sheet_names=tuple(self.workbook.sheetnames),
        )
